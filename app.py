"""
Planilla de Movilidad - METASIL S.A.C. — Web App Flask
pip install flask reportlab openpyxl
"""

import os, csv, json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    PDF_OK = True
except ImportError:
    PDF_OK = False

app = Flask(__name__)

ARCHIVO_CSV        = "registros_movilidad.csv"
ARCHIVO_XLSX       = "registros_movilidad.xlsx"
ARCHIVO_CONTADORES = "contadores_recibo.json"

# ── Colores ──
AZ    = "#1B3F6E"
AZ2   = "#2B4F82"
GOLD  = "#C8A84B"
BORDER= "#D0D7E3"
TXT   = "#1A1A2E"
TXT2  = "#5A6478"
GREEN = "#1E7F4E"


# ════════════════════════════════════════════════════════════════════════════
# Contadores
# ════════════════════════════════════════════════════════════════════════════
def cargar_contadores():
    if os.path.exists(ARCHIVO_CONTADORES):
        try:
            with open(ARCHIVO_CONTADORES, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def guardar_contadores(c):
    with open(ARCHIVO_CONTADORES, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False, indent=2)

def siguiente_recibo(dni: str) -> int:
    c = cargar_contadores()
    n = c.get(dni, 0) + 1
    c[dni] = n
    guardar_contadores(c)
    return n

def recibo_actual(dni: str) -> int:
    return cargar_contadores().get(dni, 0) + 1


# ════════════════════════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════════════════════════
def dibujar_planilla(c, datos, ox, oy, ancho, alto):
    def ln(x1,y1,x2,y2,g=0.5,col=colors.black):
        c.setStrokeColor(col); c.setLineWidth(g); c.line(x1,y1,x2,y2)

    def t(s,x,y,sz=7,bold=False,al="left",col=colors.HexColor(TXT)):
        c.setFillColor(col)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        s = str(s)
        if al=="center": c.drawCentredString(x,y,s)
        elif al=="right": c.drawRightString(x,y,s)
        else: c.drawString(x,y,s)

    yb = oy - alto
    c.setStrokeColor(colors.HexColor(AZ)); c.setLineWidth(1)
    c.rect(ox, yb, ancho, alto, fill=0, stroke=1)

    cab_h = alto * 0.13
    y_cab = oy - cab_h
    c.setFillColor(colors.white)
    c.rect(ox, y_cab, ancho, cab_h, fill=1, stroke=0)

    logo = "static/metasil_logo.png"
    logo_w = ancho * 0.22
    if os.path.exists(logo):
        c.drawImage(logo, ox+0.25*cm, y_cab+cab_h*0.1,
                    width=logo_w, height=cab_h*0.8,
                    preserveAspectRatio=True, mask="auto")

    sep1 = ox + ancho*0.26
    sep2 = ox + ancho*0.82
    for sx in (sep1, sep2):
        ln(sx, y_cab, sx, oy, g=0.7, col=colors.HexColor(AZ))

    mid_x = (sep1 + sep2) / 2
    t("PLANILLA DE MOVILIDAD", mid_x, y_cab+cab_h*0.58,
      sz=12, bold=True, al="center", col=colors.HexColor(AZ))
    t("INGENIERÍA DE PROCESOS METALÚRGICOS", mid_x, y_cab+cab_h*0.22,
      sz=6, al="center", col=colors.HexColor(TXT2))
    t("RECIBO N°", sep2+0.2*cm, y_cab+cab_h*0.65, sz=6.5, col=colors.HexColor(TXT2))
    t(str(datos.get("recibo","1")), sep2+0.2*cm, y_cab+cab_h*0.2,
      sz=14, bold=True, col=colors.HexColor(AZ))

    gold_h = alto * 0.008
    c.setFillColor(colors.HexColor(GOLD))
    c.rect(ox, y_cab-gold_h, ancho, gold_h, fill=1, stroke=0)

    dat_h = alto * 0.10
    y_dat = y_cab - gold_h - dat_h
    c.setFillColor(colors.HexColor("#EEF2FA"))
    c.rect(ox, y_dat, ancho, dat_h, fill=1, stroke=0)
    ln(ox, y_dat, ox+ancho, y_dat, g=0.5, col=colors.HexColor(BORDER))

    y1 = y_cab - gold_h - dat_h*0.32
    y2 = y_cab - gold_h - dat_h*0.72
    t("APELLIDOS Y NOMBRES:", ox+0.3*cm, y1, sz=7, bold=True, col=colors.HexColor(AZ))
    t(datos["nombre"], ox+4.5*cm, y1, sz=7.5)
    t("DNI:", ox+0.3*cm, y2, sz=7, bold=True, col=colors.HexColor(AZ))
    t(datos["dni"], ox+1.5*cm, y2, sz=7.5)
    mid = ox + ancho*0.58
    t("FECHA:", mid, y1, sz=7, bold=True, col=colors.HexColor(AZ))
    t(datos["fecha_emision"], mid+1.3*cm, y1, sz=7.5)
    t("CARGO:", mid, y2, sz=7, bold=True, col=colors.HexColor(AZ))
    t(datos["cargo"], mid+1.3*cm, y2, sz=7.5)

    cw_rel = [2.20,1.25,2.30,1.20,2.00,1.20,2.20,4.80,1.85]
    tot_rel = sum(cw_rel)
    cw = [r/tot_rel*ancho for r in cw_rel]
    def xc(i): return ox + sum(cw[:i])

    n_rows  = max(len(datos["filas"]), 3)
    fh_c1   = alto * 0.048
    fh_c2   = alto * 0.044
    fh_row  = alto * 0.062
    tabla_h = fh_c1 + fh_c2 + fh_row*n_rows
    y_tt    = y_dat
    y_tb    = y_tt - tabla_h

    c.setFillColor(colors.HexColor(AZ))
    c.rect(ox, y_tt-fh_c1, ancho, fh_c1, fill=1, stroke=0)
    for lb,i in [("N° / FECHA",0),("COD.\nCC",1),("CENTRO DE\nCOSTO",2),("DETALLE",7),("IMPORTE\nTOTAL",8)]:
        cx = xc(i)+cw[i]/2
        cy = y_tt - fh_c1/2 - 0.06*cm
        lines = lb.split("\n")
        off = (len(lines)-1)*0.11*cm/2
        for j,p in enumerate(lines):
            t(p, cx, cy+off-j*0.22*cm, sz=6, bold=True, al="center", col=colors.white)
    for lb,ci,nc in [("SALIDA",3,2),("LLEGADA",5,2)]:
        span_w = sum(cw[ci:ci+nc])
        cx = xc(ci)+span_w/2
        t(lb, cx, y_tt-fh_c1/2-0.08*cm, sz=7, bold=True, al="center", col=colors.HexColor(GOLD))

    c.setFillColor(colors.HexColor(AZ2))
    c.rect(ox, y_tt-fh_c1-fh_c2, ancho, fh_c2, fill=1, stroke=0)
    for i,lb in {3:"HORA",4:"PUNTO PARTIDA",5:"HORA",6:"PUNTO DE LLEGADA"}.items():
        t(lb, xc(i)+cw[i]/2, y_tt-fh_c1-fh_c2/2-0.08*cm,
          sz=5.5, bold=True, al="center", col=colors.HexColor("#D0E4FF"))

    y_r = y_tt - fh_c1 - fh_c2
    for idx in range(n_rows):
        bg = colors.white if idx%2==0 else colors.HexColor("#F0F5FF")
        c.setFillColor(bg)
        c.rect(ox, y_r-fh_row, ancho, fh_row, fill=1, stroke=0)
        ln(ox, y_r-fh_row, ox+ancho, y_r-fh_row, g=0.3, col=colors.HexColor(BORDER))
        for i in range(1,len(cw)):
            ln(xc(i), y_r-fh_row, xc(i), y_r, g=0.3, col=colors.HexColor(BORDER))
        if idx < len(datos["filas"]):
            f = datos["filas"][idx]
            vals = [
                f"{idx+1}  {f.get('fecha','')}",
                f.get("cod_cc",""), f.get("centro_costo",""),
                f.get("hora_salida",""), f.get("punto_partida",""),
                f.get("hora_llegada",""), f.get("punto_llegada",""),
                f.get("detalle",""), f"{float(f.get('importe',0)):.2f}"
            ]
            for i,v in enumerate(vals):
                al = "right" if i==8 else ("center" if i in(0,3,5) else "left")
                px = xc(i)+cw[i]-0.1*cm if al=="right" else \
                     xc(i)+cw[i]/2 if al=="center" else xc(i)+0.12*cm
                t(v, px, y_r-fh_row/2-0.08*cm, sz=7, al=al)
        y_r -= fh_row

    c.setStrokeColor(colors.HexColor(AZ)); c.setLineWidth(0.7)
    c.rect(ox, y_tb, ancho, tabla_h, fill=0, stroke=1)

    tot_h = alto * 0.055
    y_tot = y_tb - tot_h
    c.setFillColor(colors.HexColor("#EEF3FB"))
    c.rect(ox, y_tot, ancho, tot_h, fill=1, stroke=0)
    ln(ox, y_tot, ox+ancho, y_tot, g=0.8, col=colors.HexColor(AZ))
    total = sum(float(f.get("importe",0)) for f in datos["filas"])
    t("TOTAL S/.", ox+ancho-cw[8]-2.2*cm, y_tb-tot_h/2-0.08*cm,
      sz=7.5, bold=True, col=colors.HexColor(AZ))
    t(f"{total:.2f}", ox+ancho-0.12*cm, y_tb-tot_h/2-0.08*cm,
      sz=9, bold=True, al="right", col=colors.HexColor(GREEN))

    obs_h = alto * 0.075
    y_obs = y_tot - obs_h
    c.setFillColor(colors.HexColor("#FAFBFF"))
    c.rect(ox, y_obs, ancho, obs_h, fill=1, stroke=0)
    ln(ox, y_obs, ox+ancho, y_obs, g=0.4, col=colors.HexColor(BORDER))
    t("OBSERVACIONES: Marcar con aspa (x) el medio usado",
      ox+0.3*cm, y_tot-obs_h*0.32, sz=7, bold=True, col=colors.HexColor(AZ))
    sel = datos.get("transporte","OMNIBUS")
    tx  = ox + 0.4*cm
    for tr in ["TAXI","OMNIBUS","COLECTIVO","OTROS"]:
        c.setStrokeColor(colors.HexColor(AZ)); c.setLineWidth(0.6)
        c.rect(tx, y_tot-obs_h*0.78, 0.28*cm, 0.28*cm, fill=0)
        if sel==tr:
            t("X", tx+0.02*cm, y_tot-obs_h*0.72, sz=7.5, bold=True, col=colors.HexColor(AZ))
        t(tr, tx+0.38*cm, y_tot-obs_h*0.72, sz=7)
        tx += ancho * 0.22

    firma_h = alto * 0.22
    y_fi    = y_obs - firma_h
    fw      = ancho / 3
    ln(ox, y_fi, ox+ancho, y_fi, g=0.8, col=colors.HexColor(AZ))
    firmantes = [
        ("Autorizado por:", "Gerente Finanzas",          "Sra. Reene Moya"),
        ("Sustentado por:", "Trabajador:",                datos["nombre"]),
        ("Revisado por:",   "Asistente Administración:", ""),
    ]
    for i,(titulo,cargo_f,nombre) in enumerate(firmantes):
        fx = ox + i*fw
        if i>0:
            ln(fx, y_fi, fx, y_obs, g=0.4, col=colors.HexColor(BORDER))
        t(titulo, fx+0.3*cm, y_obs-firma_h*0.18, sz=7, bold=True, col=colors.HexColor(AZ))
        linea_y = y_obs - firma_h*0.55
        ln(fx+0.3*cm, linea_y, fx+fw-0.3*cm, linea_y, g=0.5, col=colors.HexColor("#888888"))
        t(cargo_f, fx+0.3*cm, linea_y-firma_h*0.18, sz=6.5, col=colors.HexColor(TXT2))
        if nombre:
            t(nombre, fx+0.3*cm, linea_y-firma_h*0.36, sz=6.5, bold=True)


def generar_pdf_bytes(lista_datos, por_pagina=3):
    if not PDF_OK:
        return None
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    mx, my    = 1.2*cm, 1.0*cm
    ancho     = W - 2*mx
    espacio_y = H - 2*my
    gap       = 0.35*cm
    alto_unit = (espacio_y - gap*(por_pagina-1)) / por_pagina

    idx = 0
    while idx < len(lista_datos):
        for slot in range(por_pagina):
            if idx >= len(lista_datos): break
            oy = H - my - slot*(alto_unit+gap)
            dibujar_planilla(c, lista_datos[idx], mx, oy, ancho, alto_unit)
            idx += 1
        if idx < len(lista_datos):
            c.showPage()
    c.save()
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════════════════════
# CSV / Excel
# ════════════════════════════════════════════════════════════════════════════
def guardar_csv(datos):
    hdrs = ["recibo","nombre","dni","cargo","fecha_emision","transporte",
            "fecha","cod_cc","centro_costo","hora_salida","punto_partida",
            "hora_llegada","punto_llegada","detalle","importe"]
    existe = os.path.exists(ARCHIVO_CSV)
    with open(ARCHIVO_CSV,"a",newline="",encoding="utf-8") as f:
        wr = csv.DictWriter(f, fieldnames=hdrs)
        if not existe: wr.writeheader()
        for fila in datos["filas"]:
            wr.writerow({"recibo":datos["recibo"],"nombre":datos["nombre"],
                         "dni":datos["dni"],"cargo":datos["cargo"],
                         "fecha_emision":datos["fecha_emision"],
                         "transporte":datos["transporte"],**fila})

def guardar_excel(datos):
    if not EXCEL_OK: return
    hdrs=["Recibo","Nombre","DNI","Cargo","Fecha","Transporte",
          "Fecha Mov","Cod CC","Centro","H.Sal","Partida",
          "H.Lle","Llegada","Detalle","Importe"]
    if os.path.exists(ARCHIVO_XLSX):
        wb=openpyxl.load_workbook(ARCHIVO_XLSX); ws=wb.active
    else:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Movilidad"
        for ci,h in enumerate(hdrs,1):
            cell=ws.cell(row=1,column=ci,value=h)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor="1B3F6E")
            cell.alignment=Alignment(horizontal="center")
    for fila in datos["filas"]:
        ws.append([datos["recibo"],datos["nombre"],datos["dni"],
                   datos["cargo"],datos["fecha_emision"],datos["transporte"],
                   fila.get("fecha",""),fila.get("cod_cc",""),
                   fila.get("centro_costo",""),fila.get("hora_salida",""),
                   fila.get("punto_partida",""),fila.get("hora_llegada",""),
                   fila.get("punto_llegada",""),fila.get("detalle",""),
                   float(fila.get("importe",0))])
    wb.save(ARCHIVO_XLSX)


# ════════════════════════════════════════════════════════════════════════════
# Rutas Flask
# ════════════════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/recibo_preview/<dni>")
def recibo_preview(dni):
    return jsonify({"recibo": recibo_actual(dni)})

@app.route("/generar", methods=["POST"])
def generar():
    datos = request.get_json()
    if not datos:
        return jsonify({"error": "Sin datos"}), 400

    dni        = datos.get("dni","").strip()
    num_recibo = siguiente_recibo(dni)
    datos["recibo"] = str(num_recibo)

    guardar_csv(datos)
    guardar_excel(datos)

    por_pagina = int(datos.get("por_pagina", 3))
    buf = generar_pdf_bytes([datos], por_pagina=por_pagina)
    if not buf:
        return jsonify({"error": "reportlab no instalado"}), 500

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True,
                     download_name=f"planilla_{datos['nombre']}_{ts}.pdf")


if __name__ == "__main__":
    os.makedirs("static", exist_ok=True)
    os.makedirs("templates", exist_ok=True)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)